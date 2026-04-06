#!/usr/bin/env python3
"""
Export vessels seen inside protected areas to an Excel spreadsheet.

Spatial-joins vessel positions with protected area polygons to produce
a downloadable analysis product. Each row is a unique vessel + protected
area combination with total hours, years active, and vessel metadata.

Usage:
    uv run python scripts/export_protected_area_vessels.py

Output:
    data/export/vessels_in_protected_areas.xlsx
"""

import json
import sys
from pathlib import Path

try:
    import duckdb
except ImportError:
    print("duckdb is required: uv add duckdb")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl is required: uv add openpyxl")
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent
DB_PATH = ROOT / "data" / "data.duckdb"
EXPORT_DIR = ROOT / "data" / "export"
SANCTIONS_PATH = EXPORT_DIR / "sanctioned_mmsi.json"
METADATA_PATH = EXPORT_DIR / "vessel_metadata.json"
OUTPUT_PATH = EXPORT_DIR / "vessels_in_protected_areas.xlsx"


def load_sanctions():
    if SANCTIONS_PATH.exists():
        return set(json.loads(SANCTIONS_PATH.read_text()))
    return set()


def load_metadata():
    if METADATA_PATH.exists():
        return json.loads(METADATA_PATH.read_text())
    return {}


def query_vessels_in_protected_areas(con):
    """Spatial join: vessel positions inside protected area polygons."""
    print("Running spatial join (vessel positions × protected areas)...")

    rows = con.execute("""
        SELECT
            vp.mmsi,
            vp.ship_name,
            vp.flag,
            vp.vessel_type,
            pa.area_name AS protected_area,
            pa.category,
            pa.significance,
            pa.status AS pa_status,
            ROUND(pa.area_ha / 100) AS area_km2,
            LIST(DISTINCT vp.year ORDER BY vp.year) AS years,
            COUNT(*) AS detections,
            ROUND(SUM(vp.hours), 1) AS total_hours,
            MIN(vp.entry_timestamp) AS first_seen,
            MAX(vp.exit_timestamp) AS last_seen
        FROM vessel_positions vp
        JOIN protected_areas_ocean pa
            ON ST_Contains(pa.geometry, ST_Point(vp.lon, vp.lat))
        GROUP BY
            vp.mmsi, vp.ship_name, vp.flag, vp.vessel_type,
            pa.area_name, pa.category, pa.significance, pa.status, pa.area_ha
        ORDER BY total_hours DESC
    """).fetchall()

    columns = [
        "mmsi", "ship_name", "flag", "vessel_type",
        "protected_area", "category", "significance", "pa_status", "area_km2",
        "years", "detections", "total_hours", "first_seen", "last_seen"
    ]

    print(f"  Found {len(rows)} vessel × protected area combinations")
    return columns, rows


def format_excel(wb, columns, rows, sanctions, metadata):
    """Create a formatted Excel sheet."""
    ws = wb.active
    ws.title = "Vessels in Protected Areas"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Define display columns (add enriched columns)
    display_headers = [
        "MMSI", "Vessel Name", "Flag", "Vessel Type",
        "Build Year", "DWT", "Sanctioned",
        "Protected Area", "Category", "Significance", "Status", "Area (km²)",
        "Years Active", "Detections", "Total Hours",
        "First Seen", "Last Seen"
    ]

    # Write headers
    for col_idx, header in enumerate(display_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Sanction fill
    sanction_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")

    # Write data rows
    for row_idx, row in enumerate(rows, 2):
        mmsi = str(row[0])
        meta = metadata.get(mmsi, {})
        is_sanctioned = mmsi in sanctions

        values = [
            mmsi,                                           # MMSI
            row[1] or "",                                   # Vessel Name
            row[2] or "",                                   # Flag
            row[3] or "",                                   # Vessel Type
            meta.get("y", ""),                              # Build Year
            meta.get("d", ""),                              # DWT
            "Yes" if is_sanctioned else "",                 # Sanctioned
            row[4] or "",                                   # Protected Area
            row[5] or "",                                   # Category
            row[6] or "",                                   # Significance
            row[7] or "",                                   # Status
            row[8] or "",                                   # Area km2
            ", ".join(str(y) for y in row[9]) if row[9] else "",  # Years
            row[10],                                        # Detections
            row[11],                                        # Total Hours
            row[12].strftime("%Y-%m-%d") if row[12] else "",  # First Seen
            row[13].strftime("%Y-%m-%d") if row[13] else "",  # Last Seen
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if is_sanctioned:
                cell.fill = sanction_fill

    # Auto-width columns
    for col_idx in range(1, len(display_headers) + 1):
        max_len = len(str(display_headers[col_idx - 1]))
        for row_idx in range(2, min(len(rows) + 2, 102)):  # Sample first 100 rows
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

    # Freeze top row
    ws.freeze_panes = "A2"

    # Add auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(display_headers))}{len(rows) + 1}"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.cell(row=1, column=1, value="Vessels in Protected Areas — Summary").font = Font(bold=True, size=14)
    ws2.cell(row=3, column=1, value="Total vessel × area combinations:").font = Font(bold=True)
    ws2.cell(row=3, column=2, value=len(rows))

    unique_mmsi = set(str(r[0]) for r in rows)
    ws2.cell(row=4, column=1, value="Unique vessels:").font = Font(bold=True)
    ws2.cell(row=4, column=2, value=len(unique_mmsi))

    unique_pa = set(r[4] for r in rows if r[4])
    ws2.cell(row=5, column=1, value="Protected areas with vessel activity:").font = Font(bold=True)
    ws2.cell(row=5, column=2, value=len(unique_pa))

    sanctioned_in_pa = unique_mmsi & sanctions
    ws2.cell(row=6, column=1, value="Sanctioned vessels in protected areas:").font = Font(bold=True)
    ws2.cell(row=6, column=2, value=len(sanctioned_in_pa))

    total_hours = sum(r[11] for r in rows if r[11])
    ws2.cell(row=7, column=1, value="Total vessel-hours in protected areas:").font = Font(bold=True)
    ws2.cell(row=7, column=2, value=round(total_hours, 1))

    ws2.column_dimensions["A"].width = 42
    ws2.column_dimensions["B"].width = 15


def main():
    if not DB_PATH.exists():
        print(f"Database not found: {DB_PATH}")
        print("Run 'make transform' first.")
        sys.exit(1)

    EXPORT_DIR.mkdir(parents=True, exist_ok=True)

    con = duckdb.connect(str(DB_PATH), read_only=True)
    con.execute("LOAD spatial")

    columns, rows = query_vessels_in_protected_areas(con)
    con.close()

    if not rows:
        print("No vessels found in protected areas.")
        return

    sanctions = load_sanctions()
    metadata = load_metadata()

    print("Writing Excel file...")
    wb = openpyxl.Workbook()
    format_excel(wb, columns, rows, sanctions, metadata)
    wb.save(OUTPUT_PATH)
    print(f"Exported to {OUTPUT_PATH}")
    print(f"  {len(rows)} rows across {len(set(r[4] for r in rows if r[4]))} protected areas")


if __name__ == "__main__":
    main()

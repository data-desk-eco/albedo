#!/usr/bin/env python3
"""
Ingest supplementary vessel metadata from XLSX into DuckDB.

Reads vessel metadata (build year, DWT, IMO) from an Excel file
and merges it with existing vessel data in the database.

The XLSX should have columns: mmsi, imo, build_year, dwt
(other columns are ignored)

Usage:
    uv run python scripts/ingest_vessel_metadata.py [path/to/vessels.xlsx]

Default: data/vessel_metadata.xlsx
"""

import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
DB_PATH = ROOT / "data" / "data.duckdb"
DEFAULT_XLSX = ROOT / "data" / "vessel_metadata.xlsx"


def create_dummy_xlsx(path: Path):
    """Create a dummy XLSX with sample vessel metadata for testing."""
    try:
        import openpyxl
    except ImportError:
        print("openpyxl not available, creating CSV fallback instead")
        csv_path = path.with_suffix(".csv")
        csv_path.write_text(
            "mmsi,imo,build_year,dwt\n"
            "273211040,9148580,1998,4690\n"
            "273253500,9186715,2000,5020\n"
            "273259920,9076892,1994,3780\n"
            "273295230,9251089,2003,47225\n"
            "273296810,9256423,2003,15734\n"
            "273318120,9333671,2006,70527\n"
            "273330620,9376498,2008,105313\n"
            "273345020,9448042,2010,42363\n"
            "273355640,9490503,2012,25055\n"
            "273385570,9551769,2011,18074\n"
        )
        print(f"Created dummy CSV: {csv_path}")
        return csv_path

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vessel Metadata"
    ws.append(["mmsi", "imo", "build_year", "dwt"])
    # Sample vessels (representative Arctic fleet data)
    sample_data = [
        ("273211040", "9148580", 1998, 4690),
        ("273253500", "9186715", 2000, 5020),
        ("273259920", "9076892", 1994, 3780),
        ("273295230", "9251089", 2003, 47225),
        ("273296810", "9256423", 2003, 15734),
        ("273318120", "9333671", 2006, 70527),
        ("273330620", "9376498", 2008, 105313),
        ("273345020", "9448042", 2010, 42363),
        ("273355640", "9490503", 2012, 25055),
        ("273385570", "9551769", 2011, 18074),
    ]
    for row in sample_data:
        ws.append(row)
    wb.save(path)
    print(f"Created dummy XLSX: {path} ({len(sample_data)} vessels)")
    return path


def ingest(xlsx_path: Path):
    """Ingest vessel metadata into DuckDB."""
    import duckdb

    if not DB_PATH.exists():
        print(f"Database not found at {DB_PATH}")
        print("Run 'make transform' first")
        sys.exit(1)

    con = duckdb.connect(str(DB_PATH))

    # Determine file type
    if xlsx_path.suffix == ".csv":
        read_expr = f"read_csv('{xlsx_path}')"
    else:
        # DuckDB can read xlsx with spatial extension or we use openpyxl
        # Try the spatial extension first
        try:
            con.execute("INSTALL spatial; LOAD spatial;")
            read_expr = f"st_read('{xlsx_path}')"
        except Exception:
            # Fall back to CSV
            csv_path = xlsx_path.with_suffix(".csv")
            if not csv_path.exists():
                print("Cannot read XLSX directly, converting to CSV...")
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(xlsx_path)
                    ws = wb.active
                    with open(csv_path, "w") as f:
                        for row in ws.iter_rows(values_only=True):
                            f.write(",".join(str(v) if v is not None else "" for v in row) + "\n")
                except ImportError:
                    print("openpyxl not available for XLSX conversion")
                    sys.exit(1)
            read_expr = f"read_csv('{csv_path}')"

    # Create or replace vessel_metadata table
    con.execute(f"""
        CREATE OR REPLACE TABLE vessel_metadata AS
        SELECT
            CAST(mmsi AS VARCHAR) as mmsi,
            CAST(imo AS VARCHAR) as imo,
            CAST(build_year AS INTEGER) as build_year,
            CAST(dwt AS DOUBLE) as dwt
        FROM {read_expr}
        WHERE mmsi IS NOT NULL
    """)

    count = con.execute("SELECT COUNT(*) FROM vessel_metadata").fetchone()[0]
    print(f"Loaded {count} vessel metadata records into vessel_metadata table")

    # Show matched vessels
    matched = con.execute("""
        SELECT COUNT(DISTINCT va.mmsi)
        FROM vessel_activity va
        JOIN vessel_metadata vm ON va.mmsi = vm.mmsi
    """).fetchone()[0]
    print(f"  {matched} matched with existing vessel_activity data")

    # Export vessel_metadata.json for frontend tooltips
    rows = con.execute("""
        SELECT mmsi, imo, build_year, CAST(dwt AS INTEGER) as dwt
        FROM vessel_metadata
        WHERE mmsi IS NOT NULL
    """).fetchall()
    meta = {}
    for mmsi, imo, build_year, dwt in rows:
        entry = {}
        if imo:
            entry["imo"] = imo
        if build_year:
            entry["y"] = build_year
        if dwt:
            entry["d"] = dwt
        if entry:
            meta[mmsi] = entry

    export_path = ROOT / "data" / "export" / "vessel_metadata.json"
    export_path.write_text(json.dumps(meta, separators=(",", ":")))
    print(f"  Exported {len(meta)} entries to {export_path}")

    con.close()


if __name__ == "__main__":
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX

    if not xlsx_path.exists():
        print(f"Metadata file not found: {xlsx_path}")
        print("Creating dummy data for testing...")
        xlsx_path = create_dummy_xlsx(xlsx_path)

    ingest(xlsx_path)
